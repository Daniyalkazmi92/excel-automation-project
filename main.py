import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, numbers
from openpyxl.utils import get_column_letter
import os

# Step 1: Ask user for file name
input_file = input("Enter the name of your Excel file (e.g., sales_data.xlsx): ")

# Step 2: File check
if not os.path.exists(input_file):
    print("❌ File not found! Please check the file name.")
    exit()

# Step 3: Load file and validate columns
try:
    df = pd.read_excel(input_file)
except Exception as e:
    print("❌ Error reading Excel file:", str(e))
    exit()

required_columns = {"Date", "Product", "Quantity", "Price"}
if not required_columns.issubset(df.columns):
    print("❌ Missing required columns. Required:", required_columns)
    print("Your file has columns:", df.columns.tolist())
    exit()

# Step 4: Perform calculations
df["Total"] = df["Quantity"] * df["Price"]

# Step 5: Create summary
grand_total = df["Total"].sum()
top_product = df.groupby("Product")["Total"].sum().idxmax()
top_product_sales = df.groupby("Product")["Total"].sum().max()

summary_df = pd.DataFrame({
    "Metric": ["Grand Total Sales", "Top Product", "Top Product Sales"],
    "Value": [grand_total, top_product, top_product_sales]
})

# Step 6: Save to output file
output_file = "sales_report.xlsx"
with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    df.to_excel(writer, index=False, sheet_name="Report")
    summary_df.to_excel(writer, index=False, sheet_name="Summary")

# Step 7: Formatting
wb = load_workbook(output_file)
ws = wb["Report"]

# Bold headers + center
for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")

# Currency format
for col_letter in ["D", "E"]:  # Price and Total
    for row in range(2, ws.max_row + 1):
        ws[f"{col_letter}{row}"].number_format = '"Rs"#,##0.00'

# Auto column width
for col in ws.columns:
    max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
    ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

wb.save(output_file)
print("✅ Success! Report saved as:", output_file)
